import openpyxl
import json


corba = []
ara_yemek = []
new_corba = []
ana_etli_yemek = []
ana_etsiz_yemek = []

wb = openpyxl.load_workbook('yemek.xlsx')
ws = wb.active
#print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
number=13
column= chr(65)
for i in range(1,8):
    for j in range(1,6):
        corba.append(ws[str(column)+str(number)].value)
        number=number+9
    number=13    
    column = chr(ord(column)+1)
#print(corba)    


#############################################################################################################################
number2=16
column2= chr(65)
flag =0  
for k in range(1,8):
    for l in range(1,30):
        ara_yemek.append(ws[str(column2)+str(number2)].value)
        if(flag <4):
            number2=number2+1
            flag=flag+1
        else:
           number2=number2+5
           flag=0
    number2=16
    column2 = chr(ord(column2)+1)
    flag=0       
    
#print(ara_yemek)

number3=14
column3= chr(65)
for i in range(1,8):
    for j in range(1,6):
        ana_etli_yemek.append(ws[str(column3)+str(number3)].value)
        number3=number3 + 9
    number3=14    
    column3= chr(ord(column3)+1)


#print(ana_etli_yemek)

new_ana_etli_yemek = []
for ana_etli_yemek_elemani in ana_etli_yemek:
     if ana_etli_yemek_elemani is not None:
         ana_etli_yemekler = ana_etli_yemek_elemani.split('-')
         new_ana_etli_yemek.extend([ana_etli_yemek.strip() for ana_etli_yemek in ana_etli_yemekler])
     else:
         print("None değeri ile karşılaşıldı, bu öğe atlandı.")

#print(new_ana_etli_yemek)
#print(ana_etsiz_yemek)

number4=15
column4= chr(65)
for i in range(1,8):
    for j in range(1,6):
        ana_etsiz_yemek.append(ws[str(column4)+str(number4)].value)
        number4=number4 + 9
    number4=15    
    column4= chr(ord(column4)+1)

# print(ana_etsiz_yemek)


new_ana_etsiz_yemek = []
for ana_etsiz_yemek_elemani in ana_etsiz_yemek:
     if ana_etsiz_yemek_elemani is not None:
        ana_etsiz_yemekler = ana_etsiz_yemek_elemani.split('-')
        new_ana_etsiz_yemek.extend([ana_etsiz_yemek.strip() for ana_etsiz_yemek in ana_etsiz_yemekler])
     else:
         print("None değeri ile karşılaşıldı, bu öğe atlandı.")

#print(new_ana_etsiz_yemek)



# for corba_elemani in corba:
#     if corba_elemani is not None:
#         corbalar = corba_elemani.split('-')
#         new_corba.extend([corba.strip() for corba in corbalar])
#     else:
#         print("None değeri ile karşılaşıldı, bu öğe atlandı.")

# print(new_corba)        

        ###################################################################################################
new_ara_yemek = []
for ara_yemek_elemani in ara_yemek:
    if ara_yemek_elemani is not None:
        corbalar = ara_yemek_elemani.split('-')
        new_ara_yemek.extend([corba.strip() for corba in corbalar])
    else:
        print("None değeri ile karşılaşıldı, bu öğe atlandı.")

#print(new_ara_yemek)     

### CONVERT JSON ##############
corba_dict_list = []
unique_corba = set(new_corba)
# for corba_elemani in unique_corba:
#     parts = corba_elemani.split('(')
#     corba_ismi = parts[0].strip()  # Çorba ismi
#     kalori = parts[1].replace(')', '').strip()  # Kalori bilgisi
    
#     corba_dict = {
#         "isim": corba_ismi,
#         "kalori": kalori,
#         "tür": "corba"
#     }
    
#     corba_dict_list.append(corba_dict)

# corba_json = json.dumps(corba_dict_list, indent=4, ensure_ascii=False)
# print(corba_json)

unique_ara_yemek = set(new_ara_yemek)

ara_yemek_dict_list = []

for yemek_elemani in unique_ara_yemek:
    parts = yemek_elemani.split('(')
    
    if len(parts) > 1:  
        yemek_ismi = parts[0].strip()  
        kalori = parts[1].replace(')', '').strip()  
    else:
        yemek_ismi = yemek_elemani.strip()
        kalori = "Bilgi yok"
    
    yemek_dict = {
        "isim": yemek_ismi,
        "kalori": kalori,
        "tur": "Yardımcı Yemek"
    }
    
    ara_yemek_dict_list.append(yemek_dict)

ara_yemek_json = json.dumps(ara_yemek_dict_list, indent=4, ensure_ascii=False)
#print(ara_yemek_json)


unique_ana_etli_yemek = set(new_ana_etli_yemek)

ana_etli_yemek_dict_list = []

for yemek_elemani in unique_ana_etli_yemek:
    parts = yemek_elemani.split('(')
    
    if len(parts) > 1:  
        yemek_ismi = parts[0].strip()  
        kalori = parts[1].replace(')', '').strip()  
    else:
        yemek_ismi = yemek_elemani.strip()
        kalori = "Bilgi yok"
    
    yemek_dict = {
        "isim": yemek_ismi,
        "kalori": kalori,
        "tur": "Etli Yemek"
    }
    
    ana_etli_yemek_dict_list.append(yemek_dict)

ana_etli_yemek_json = json.dumps(ana_etli_yemek_dict_list, indent=4, ensure_ascii=False)



#print(ana_etsiz_yemek_json)

#print(unique_ana_etsiz_yemek)
#print(ana_etli_yemek_json)


unique_ana_etsiz_yemek = set(new_ana_etsiz_yemek)

ana_etsiz_yemek_dict_list = []

for yemek_elemani in unique_ana_etsiz_yemek:
    parts = yemek_elemani.split('(')
    
    if len(parts) > 1:  
        yemek_ismi = parts[0].strip()  
        kalori = parts[1].replace(')', '').strip()  
    else:
        yemek_ismi = yemek_elemani.strip()
        kalori = "Bilgi yok"
    
    yemek_dict = {
        "isim": yemek_ismi,
        "kalori": kalori,
        "tur": "Etsiz Yemek"
    }
    ana_etsiz_yemek_dict_list.append(yemek_dict)

ana_etsiz_yemek_json = json.dumps(ana_etsiz_yemek_dict_list, indent=4, ensure_ascii=False)

print(unique_ana_etli_yemek)
print("-------------------------*********************---------------------------------")

print("++++++++++++++++++++++++++++++1++++++++++++++++++++++++++++++++++++++++")
print(unique_ana_etsiz_yemek)
print("++++++++++++++++++++++++++++++2++++++++++++++++++++++++++++++++++++++++")
print(new_ana_etsiz_yemek)
print("+++++++++++++++++++++++++++++++3+++++++++++++++++++++++++++++++++++++++")
print(ana_etsiz_yemek_dict_list)
print("++++++++++++++++++++++++++++++++4++++++++++++++++++++++++++++++++++++++")
print(ana_etsiz_yemek_json)
print("+++++++++++++++++++++++++++++++++5+++++++++++++++++++++++++++++++++++++")
